import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Documents/Desktop/bbinge_fc_goat_photo_targets_136.xlsx"
OUTPUT = ROOT / "src/data/player-photo-audit.json"

players_data = json.loads((ROOT / "src/data/goat-players.json").read_text(encoding="utf-8"))
players_by_original = {player["nameOriginal"]: player for player in players_data["players"]}
players_by_id = {player["id"]: player for player in players_data["players"]}
aliases = {
    "Lev Yashin": "yashin",
    "Johan Cruyff": "cruyff",
    "Oleh Blokhin": "blokhin",
    "Bum-kun Cha": "cha-bum-kun",
    "Andriy Shevchenko": "shevchenko",
    "Ji-sung Park": "park-ji-sung",
    "Mohamed Salah": "salah",
    "Heung-min Son": "son-heung-min",
    "Min-jae Kim": "kim-min-jae",
}

sheet = load_workbook(WORKBOOK, data_only=True)["사진 타깃표"]
rows = []
for row_number in range(2, sheet.max_row + 1):
    (
        _, era, player_kr, player_original, target_years, preferred_clubs,
        secondary_allowed, excluded, common_criteria, _, source_url,
        license_info, notes,
    ) = [sheet.cell(row_number, column).value for column in range(1, 14)]
    player = players_by_original.get(player_original) or players_by_id.get(aliases.get(player_original, ""))
    if not player:
        raise ValueError(f"GOAT ID를 찾을 수 없음: {player_kr} / {player_original}")
    rows.append({
        "id": player["id"],
        "player": player_original,
        "playerKr": player_kr,
        "era": era,
        "targetYears": target_years,
        "preferredClubs": [value.strip() for value in preferred_clubs.split("/") if value.strip()] if preferred_clubs else [],
        "secondaryAllowed": [secondary_allowed] if secondary_allowed else [],
        "excluded": [value.strip() for value in excluded.split("·") if value.strip()] if excluded else [],
        "commonCriteria": common_criteria,
        "sourceUrl": source_url or "",
        "license": license_info or "",
        "status": "missing",
        "notes": notes or "",
        "photoPosition": {
            "desktop": "50% 22%", "laptop": "50% 24%",
            "tablet": "50% 26%", "mobile": "50% 28%",
        },
        "photoScale": {"desktop": 1.0, "laptop": 1.0, "tablet": 1.0, "mobile": 1.0},
        "review": {
            "identityConfirmed": False, "eraConfirmed": False,
            "clubOrNationalTeamConfirmed": False, "licenseConfirmed": False,
            "desktopChecked": False, "laptopChecked": False,
            "tabletChecked": False, "mobileChecked": False,
        },
    })

if len(rows) != 136 or len({row["id"] for row in rows}) != 136:
    raise ValueError(f"136명 고유 ID 검증 실패: rows={len(rows)}, ids={len({row['id'] for row in rows})}")

OUTPUT.write_text(json.dumps({
    "version": "1.0",
    "sourceWorkbook": WORKBOOK.name,
    "playerCount": len(rows),
    "players": rows,
}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(f"생성 완료: {OUTPUT} ({len(rows)}명)")
